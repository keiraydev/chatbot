import openpyxl # openpyxl 모듈 불러오기

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '회원정보'

# 표 헤더 컬럼 저장
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

# 표 내용 저장
members = [('kei', '010-1234-1234'), ('hong', '010-4321-1234')]
row_num = 2 # 1번째 row는 타이틀 위치

for r, member in enumerate(members): # 회원정보 목록 탐색
    for c, v in enumerate(member): # 이름, 전화번호 컬럼 탐색
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('./members.xlsx')
wb.close()