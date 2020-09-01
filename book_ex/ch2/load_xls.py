import openpyxl # openpyxl 모듈 불러오기

wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['Sheet1']
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=2, column=1).value)
wb.close()